# reporter.py
import logging
import io
import pandas as pd
from datetime import datetime
from decimal import Decimal

# ایمپورت‌های تلگرام (سازگار با v20+)
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes
)

# ایمپورت‌های استایل‌دهی برای اکسل
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# فایل‌های کانفیگ و دیتابیس
import config
import db_utils

# تنظیمات لاگ‌گیری
logging.basicConfig(level=config.BOT["LOG_LEVEL"], format='%(asctime)s - %(threadName)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# متن دکمه
REPORT_BUTTON_TEXT = "📊 دریافت گزارش مالی"

# --- تعریف استایل‌های اکسل ---
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="B Nazanin")
PROFIT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PROFIT_FONT = Font(color="006100", name="B Nazanin")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LOSS_FONT = Font(color="9C0006", name="B Nazanin")
TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTAL_FONT = Font(bold=True, name="B Nazanin")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# --- دکوراتور ادمین (بدون تغییر) ---
def restricted(func):
    """دسترسی به دستور را فقط به ادمین تعریف شده در کانفیگ محدود می‌کند."""
    async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        user_id = update.effective_user.id
        if user_id != config.TELEGRAM["ADMIN_CHAT_ID"]:
            logger.warning(f"دسترسی غیرمجاز به ربات گزارش‌گیر از ID: {user_id}")
            await update.message.reply_text("شما مجاز به استفاده از این ربات نیستید.")
            return
        return await func(update, context, *args, **kwargs)
    return wrapped

# --- بخش محاسبات و دیتابیس (ارتقا یافته) ---

def calculate_profit(row):
    """سود دقیق را بر اساس قیمت‌های سفارش‌گذاری محاسبه می‌کند."""
    try:
        gross_buy_qty = Decimal(row['buy_executed_quantity']) + Decimal(row['buy_fee'])
        cost_tmn = gross_buy_qty * Decimal(row['entry_price'])
        revenue_tmn = Decimal(row['sell_executed_quantity']) * Decimal(row['exit_price'])
        sell_fee_tmn = Decimal(row['sell_fee'])
        profit = revenue_tmn - cost_tmn - sell_fee_tmn
        
        # سه ستون جدید برای تحلیل اضافه می‌کنیم
        row['cost_tmn'] = cost_tmn
        row['revenue_tmn'] = revenue_tmn
        row['profit_tmn'] = profit
        return row
        
    except Exception:
        return None

def fetch_report_data():
    """
    داده‌های حسابرسی را استخراج، پردازش و برای تجمیع آماده می‌کند.
    """
    query = "SELECT * FROM trade_signals"
    all_trades = db_utils.query_db(query, fetch='all')
    
    if not all_trades:
        return None
        
    df = pd.DataFrame(all_trades)
    
    # 1. سود/زیان تحقق یافته (محاسبه فردی)
    completed_trades_df = df[df['status'] == 'SELL_ORDER_FILLED'].copy()
    if not completed_trades_df.empty:
        completed_trades_df = completed_trades_df.apply(calculate_profit, axis=1)
        total_realized_profit = completed_trades_df['profit_tmn'].sum()
    else:
        total_realized_profit = 0

    # 2. پوزیشن‌های باز (محاسبه فردی)
    open_positions_df = df[df['status'].isin(['BUY_ORDER_FILLED', 'SELL_ORDER_PLACED'])].copy()
    if not open_positions_df.empty:
        open_positions_df['cost_tmn'] = (
            pd.to_numeric(open_positions_df['buy_executed_quantity']) + 
            pd.to_numeric(open_positions_df['buy_fee'])
        ) * pd.to_numeric(open_positions_df['entry_price'])
        total_invested_in_open = open_positions_df['cost_tmn'].sum()
    else:
        total_invested_in_open = 0

    # 3. آمار کلی (فارسی)
    summary_data = {
        "پارامتر": [
            "کل سود خالص تحقق یافته (تومان)",
            "تعداد کل معاملات آغاز شده",
            "تعداد معاملات موفق (بسته شده)",
            "تعداد پوزیشن‌های باز (در حال معامله)",
            "تعداد معاملات لغو شده (تایم‌اوت)",
            "تعداد معاملات دچار خطا"
        ],
        "مقدار": [
            f"{total_realized_profit:,.0f}",
            len(df),
            len(completed_trades_df),
            len(open_positions_df),
            len(df[df['status'] == 'CANCELED_TIMEOUT']),
            len(df[df['status'] == 'ERROR'])
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # 4. --- تجمیع سود بر اساس دارایی (Aggregation) ---
    if not completed_trades_df.empty:
        aggregated_profit = completed_trades_df.groupby('asset_name').agg(
            total_cost_tmn=('cost_tmn', 'sum'),
            total_revenue_tmn=('revenue_tmn', 'sum'),
            total_profit_tmn=('profit_tmn', 'sum'),
            trade_count=('id', 'count')
        ).reset_index()
    else:
        aggregated_profit = pd.DataFrame(columns=['asset_name', 'total_cost_tmn', 'total_revenue_tmn', 'total_profit_tmn', 'trade_count'])

    # 5. --- تجمیع پوزیشن‌های باز بر اساس دارایی (Aggregation) ---
    if not open_positions_df.empty:
        aggregated_open = open_positions_df.groupby('asset_name').agg(
            total_invested_tmn=('cost_tmn', 'sum'),
            total_quantity=('buy_executed_quantity', 'sum'),
            position_count=('id', 'count')
        ).reset_index()
    else:
        aggregated_open = pd.DataFrame(columns=['asset_name', 'total_invested_tmn', 'total_quantity', 'position_count'])

    return summary_df, completed_trades_df, aggregated_profit, open_positions_df, aggregated_open

def apply_styles_to_sheet(ws, df, number_format="#,##0"):
    """
    استایل‌های هدر، ستون‌ها و رنگ‌بندی شرطی را اعمال می‌کند.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # 1. اعمال استایل هدر
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # 2. اعمال استایل و فرمت به سلول‌های داده
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.font = Font(name="B Nazanin")
            
            # فرمت اعداد و چینش
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = number_format
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_CENTER
            
            # 3. رنگ‌بندی شرطی (مخصوص ستون 'سود خالص')
            header_name = ws.cell(row=1, column=col).value
            if header_name in ["سود خالص (تومان)", "مقدار"] and isinstance(cell.value, (int, float, Decimal)):
                if cell.value > 0:
                    cell.fill = PROFIT_FILL
                    cell.font = PROFIT_FONT
                elif cell.value < 0:
                    cell.fill = LOSS_FILL
                    cell.font = LOSS_FONT

    # 4. تنظیم عرض ستون‌ها
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].best_fit = True
        ws.column_dimensions[column_letter].width = ws.column_dimensions[column_letter].width + 5 # کمی فضای اضافه

    # 5. تنظیمات راست به چپ (RTL)
    ws.sheet_view.rightToLeft = True

def dataframe_to_styled_sheet(ws, df, persian_headers):
    """DataFrame را با هدرهای فارسی به شیت اکسل منتقل می‌کند"""
    # ابتدا هدرهای فارسی را می‌نویسیم
    ws.append(persian_headers)
    
    # سپس داده‌ها را (بدون هدر اصلی) اضافه می‌کنیم
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        ws.append(row)
    
    # استایل‌ها را اعمال می‌کنیم
    apply_styles_to_sheet(ws, df)


def create_styled_excel_report(report_data):
    """
    گزارش اکسل حرفه‌ای، فارسی و رنگ‌بندی شده ایجاد می‌کند.
    """
    summary_df, completed_trades_df, aggregated_profit, open_positions_df, aggregated_open = report_data
    
    output_buffer = io.BytesIO()
    wb = Workbook() # یک ورک‌بوک جدید می‌سازیم
    
    # --- شیت ۱: خلاصه عملکرد ---
    ws_summary = wb.active
    ws_summary.title = "خلاصه عملکرد"
    dataframe_to_styled_sheet(ws_summary, summary_df, ["پارامتر", "مقدار"])
    apply_styles_to_sheet(ws_summary, summary_df, number_format="General") # فرمت خاصی برای این شیت نیاز نیست

    # --- شیت ۲: سود تجمیعی (محقق شده) ---
    if not aggregated_profit.empty:
        ws_agg_profit = wb.create_sheet(title="سود تجمیعی (محقق شده)")
        headers = ["نام دارایی", "کل هزینه (تومان)", "کل درآمد (تومان)", "سود خالص (تومان)", "تعداد معاملات"]
        dataframe_to_styled_sheet(ws_agg_profit, aggregated_profit, headers)

    # --- شیت ۳: پوزیشن‌های باز (تجمیعی) ---
    if not aggregated_open.empty:
        ws_agg_open = wb.create_sheet(title="پوزیشن‌های باز (تجمیعی)")
        headers = ["نام دارایی", "کل سرمایه درگیر (تومان)", "مقدار کل دارایی", "تعداد پوزیشن"]
        dataframe_to_styled_sheet(ws_agg_open, aggregated_open, headers)

    # --- شیت ۴: جزئیات معاملات بسته شده ---
    if not completed_trades_df.empty:
        ws_details_comp = wb.create_sheet(title="جزئیات معاملات بسته شده")
        headers = [
            "ID", "دارایی", "جفت‌ارز", "قیمت ورود", "قیمت خروج", "استراتژی", 
            "وضعیت", "ID خرید", "مقدار خام", "مقدار فرمت‌شده", 
            "مقدار خالص خرید", "کارمزد خرید (ارز)", "ID فروش", "مقدار فروش", "کارمزد فروش (تومان)", 
            "یادداشت", "تاریخ ایجاد", "تاریخ آپدیت", 
            "هزینه (تومان)", "درآمد (تومان)", "سود خالص (تومان)"
        ]
        # ستون‌های مورد نیاز را انتخاب می‌کنیم
        cols_to_show = [
            'id', 'asset_name', 'pair', 'entry_price', 'exit_price', 'strategy_name', 
            'status', 'buy_client_order_id', 'buy_quantity_raw', 'buy_quantity_formatted', 
            'buy_executed_quantity', 'buy_fee', 'sell_client_order_id', 'sell_executed_quantity', 'sell_fee',
            'notes', 'created_at', 'updated_at', 
            'cost_tmn', 'revenue_tmn', 'profit_tmn'
        ]
        dataframe_to_styled_sheet(ws_details_comp, completed_trades_df[cols_to_show], headers)

    # --- شیت ۵: جزئیات پوزیشن‌های باز ---
    if not open_positions_df.empty:
        ws_details_open = wb.create_sheet(title="جزئیات پوزیشن‌های باز")
        headers = [
            "ID", "دارایی", "جفت‌ارز", "قیمت ورود", "قیمت خروج", "استراتژی", 
            "وضعیت", "ID خرید", "مقدار خالص خرید", "کارمزد خرید (ارز)", "تاریخ ایجاد", 
            "سرمایه درگیر (تومان)"
        ]
        cols_to_show = [
            'id', 'asset_name', 'pair', 'entry_price', 'exit_price', 'strategy_name', 
            'status', 'buy_client_order_id', 'buy_executed_quantity', 'buy_fee', 'created_at',
            'cost_tmn'
        ]
        dataframe_to_styled_sheet(ws_details_open, open_positions_df[cols_to_show], headers)
    
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# --- توابع ربات تلگرام (سازگار با v20+) ---

@restricted
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """تابع /start. پیام خوشامد و دکمه گزارش را ارسال می‌کند."""
    keyboard = [
        [KeyboardButton(REPORT_BUTTON_TEXT)]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        "سلام ادمین! برای دریافت گزارش مالی روی دکمه زیر کلیک کنید:",
        reply_markup=reply_markup
    )

@restricted
async def send_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    هندلر دکمه گزارش. گزارش را تولید و ارسال می‌کند.
    """
    user_name = update.effective_user.first_name
    logger.info(f"درخواست گزارش از کاربر {user_name} (Admin) دریافت شد...")
    await update.message.reply_text("در حال تهیه گزارش مالی پیشرفته... لطفاً چند لحظه صبر کنید.")

    try:
        # 1. دریافت داده‌های پردازش شده
        report_data = fetch_report_data()
        
        if not report_data:
            await update.message.reply_text("هیچ داده‌ای برای گزارش یافت نشد.")
            return
            
        # 2. ساخت فایل اکسل استایل‌دار
        excel_file_buffer = create_styled_excel_report(report_data)
        
        filename = f"Financial_Report_Styled_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        
        # 3. ارسال فایل
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=excel_file_buffer,
            filename=filename,
            caption="گزارش عملکرد مالی (فارسی و تجمیعی)"
        )
        logger.info(f"گزارش با موفقیت برای {user_name} ارسال شد.")

    except Exception as e:
        logger.error(f"خطا در ایجاد یا ارسال گزارش: {e}", exc_info=True)
        await update.message.reply_text(f"خطا در پردازش گزارش: {e}")

# --- تابع اصلی (سازگار با v20+) ---
def main():
    """
    حلقه اصلی ماژول گزارش‌گیری (اجرای ربات تلگرام با سینتکس v20+).
    """
    if not config.TELEGRAM.get("BOT_TOKEN"):
        logger.critical("توکن ربات تلگرام در کانفیگ تنظیم نشده است. ماژول گزارش‌گیری غیرفعال شد.")
        return
    if not config.TELEGRAM.get("ADMIN_CHAT_ID"):
        logger.critical("ADMIN_CHAT_ID در کانفیگ تنظیم نشده است. ماژول گزارش‌گیری غیرفعال شد.")
        return

    # ساخت Application با استفاده از Builder
    application = Application.builder().token(config.TELEGRAM["BOT_TOKEN"]).build()

    # اضافه کردن هندلرها
    application.add_handler(CommandHandler("start", start))
    
    # فیلتر کردن بر اساس متن دکمه (سازگار با v20+)
    application.add_handler(MessageHandler(
        filters.TEXT & filters.Regex(f"^{REPORT_BUTTON_TEXT}$"), 
        send_report
    ))

    logger.info("ماژول گزارش‌گیری (Reporter) شروع به کار کرد و منتظر دستورات تلگرام است...")
    
    # اجرای ربات
    application.run_polling()


if __name__ == "__main__":
    main()
